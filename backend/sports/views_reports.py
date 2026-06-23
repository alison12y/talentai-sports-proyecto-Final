import openpyxl
from openpyxl.styles import Font, PatternFill
from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.utils.dateparse import parse_date
from .models import Partido, Asistencia, EstadisticaPartido, EvolucionFisica, Equipo, Jugador, JugadorEquipo
from .serializers import (
    PartidoSerializer, AsistenciaSerializer, EstadisticaPartidoSerializer, 
    EvolucionFisicaSerializer, EquipoSerializer, JugadorSerializer
)

from django.utils import timezone

def build_report_data(request):
    tipo = request.query_params.get('tipo')
    equipo_id = request.query_params.get('equipo')
    jugador_id = request.query_params.get('jugador')
    fecha_inicio = request.query_params.get('fecha_inicio')
    fecha_fin = request.query_params.get('fecha_fin')

    if not tipo:
        return {
            "resumen": {},
            "partidos": [],
            "asistencias": [],
            "estadisticas": [],
            "evolucion_fisica": [],
            "alertas": [],
            "mensaje": "Reporte generado correctamente"
        }, None

    if tipo == 'jugador' and not jugador_id:
        return None, "Debes seleccionar un jugador."
    if tipo == 'equipo' and not equipo_id:
        return None, "Debes seleccionar un equipo."

    # Filter base
    q_partidos = Partido.objects.all()
    q_asistencias = Asistencia.objects.all()
    q_estadisticas = EstadisticaPartido.objects.all()
    q_evolucion = EvolucionFisica.objects.all()

    if fecha_inicio:
        q_partidos = q_partidos.filter(fecha_inicio__date__gte=parse_date(fecha_inicio))
        q_asistencias = q_asistencias.filter(fecha__gte=parse_date(fecha_inicio))
        q_estadisticas = q_estadisticas.filter(partido__fecha_inicio__date__gte=parse_date(fecha_inicio))
        q_evolucion = q_evolucion.filter(fecha_medicion__gte=parse_date(fecha_inicio))
    
    if fecha_fin:
        q_partidos = q_partidos.filter(fecha_inicio__date__lte=parse_date(fecha_fin))
        q_asistencias = q_asistencias.filter(fecha__lte=parse_date(fecha_fin))
        q_estadisticas = q_estadisticas.filter(partido__fecha_inicio__date__lte=parse_date(fecha_fin))
        q_evolucion = q_evolucion.filter(fecha_medicion__lte=parse_date(fecha_fin))

    resumen = {}
    if tipo == 'jugador':
        jugador = Jugador.objects.filter(id=jugador_id).first()
        if not jugador:
            return None, "Jugador no encontrado."
        q_asistencias = q_asistencias.filter(jugador=jugador)
        q_estadisticas = q_estadisticas.filter(jugador=jugador)
        q_evolucion = q_evolucion.filter(jugador=jugador)
        
        partidos_ids = set(q_estadisticas.values_list('partido_id', flat=True))
        q_partidos = q_partidos.filter(id__in=partidos_ids)

        jugador_equipo = JugadorEquipo.objects.filter(jugador=jugador, activo=True).first()

        resumen = {
            "nombre": f"{jugador.nombre} {jugador.apellido}",
            "equipo": jugador_equipo.equipo.nombre if jugador_equipo and jugador_equipo.equipo else "Sin equipo",
            "posicion": getattr(jugador, 'posicion_principal', "N/A")
        }
    elif tipo == 'equipo':
        equipo = Equipo.objects.filter(id=equipo_id).first()
        if not equipo:
            return None, "Equipo no encontrado."
        q_partidos = q_partidos.filter(equipo=equipo)
        
        jugador_ids = JugadorEquipo.objects.filter(equipo=equipo, activo=True).values_list('jugador_id', flat=True)
        jugadores_equipo = Jugador.objects.filter(id__in=jugador_ids)
        
        q_asistencias = q_asistencias.filter(jugador__in=jugadores_equipo)
        q_estadisticas = q_estadisticas.filter(jugador__in=jugadores_equipo)
        q_evolucion = q_evolucion.filter(jugador__in=jugadores_equipo)
        
        # Determine if we have any stats
        # Ensure we return valid lists even if no data
        resumen = {
            "nombre": equipo.nombre,
            "categoria": equipo.categoria.nombre if equipo.categoria else "",
            "total_jugadores": jugadores_equipo.count()
        }

    partidos_data = PartidoSerializer(q_partidos.distinct(), many=True).data
    asistencias_data = AsistenciaSerializer(q_asistencias.distinct(), many=True).data
    estadisticas_data = EstadisticaPartidoSerializer(q_estadisticas.distinct(), many=True).data
    evolucion_data = EvolucionFisicaSerializer(q_evolucion.distinct(), many=True).data

    # Enrich with jugador_nombre_completo to avoid showing UUIDs
    # We will gather all unique player IDs from the data
    jugador_ids_all = set()
    for lista in [asistencias_data, estadisticas_data, evolucion_data]:
        for item in lista:
            if 'jugador' in item:
                jugador_ids_all.add(item['jugador'])
                
    jugadores_dict = {
        j.id: f"{j.nombre} {j.apellido}" 
        for j in Jugador.objects.filter(id__in=jugador_ids_all)
    }
    
    for lista in [asistencias_data, estadisticas_data, evolucion_data]:
        for item in lista:
            if 'jugador' in item:
                item['jugador_nombre_completo'] = jugadores_dict.get(item['jugador'], "Desconocido")

    return {
        "resumen": resumen,
        "partidos": partidos_data,
        "asistencias": asistencias_data,
        "estadisticas": estadisticas_data,
        "evolucion_fisica": evolucion_data,
        "alertas": []
    }, None

class ReporteRendimientoView(APIView):
    def get(self, request):
        data, error = build_report_data(request)
        if error:
            return Response({"error": error}, status=status.HTTP_400_BAD_REQUEST)
        return Response(data)

class ReporteRendimientoExcelView(APIView):
    def get(self, request):
        data, error = build_report_data(request)
        if error:
            return Response({"error": error}, status=status.HTTP_400_BAD_REQUEST)

        wb = openpyxl.Workbook()
        
        def style_header(sheet, headers):
            sheet.append(headers)
            for col_idx in range(1, len(headers) + 1):
                cell = sheet.cell(row=1, column=col_idx)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="1E293B", fill_type="solid")
        
        # Hoja 1: Resumen
        ws_resumen = wb.active
        ws_resumen.title = "Resumen"
        resumen = data["resumen"]
        tipo_rep = request.query_params.get('tipo', 'Desconocido')
        f_inicio = request.query_params.get('fecha_inicio', 'Todas')
        f_fin = request.query_params.get('fecha_fin', 'Todas')

        style_header(ws_resumen, ["Campo", "Valor"])
        ws_resumen.append(["Tipo de reporte", tipo_rep.capitalize()])
        ws_resumen.append(["Nombre", resumen.get('nombre', 'N/A')])
        # We don't have direct access to club here easily without fetching, just say N/A if not passed.
        ws_resumen.append(["Club", "N/A"])
        ws_resumen.append(["Fecha de generación", timezone.now().strftime("%Y-%m-%d %H:%M")])
        ws_resumen.append(["Filtros aplicados", f"Inicio: {f_inicio} - Fin: {f_fin}"])
        ws_resumen.append(["Total partidos", len(data["partidos"])])
        ws_resumen.append(["Total asistencias", len(data["asistencias"])])
        ws_resumen.append(["Total estadísticas", len(data["estadisticas"])])
        ws_resumen.append(["Total registros físicos", len(data["evolucion_fisica"])])
            
        # Hoja 2: Partidos
        ws_partidos = wb.create_sheet("Partidos")
        if not data["partidos"]:
            ws_partidos.append(["Sin datos registrados."])
        else:
            style_header(ws_partidos, ["Fecha", "Equipo", "Rival", "Resultado", "Goles local", "Goles rival", "Ubicación"])
            for p in data["partidos"]:
                ws_partidos.append([
                    str(p.get('fecha', p.get('fecha_inicio', 'N/A'))), 
                    str(p.get('equipo', 'N/A')), 
                    str(p.get('nombre_rival', 'N/A')), 
                    str(p.get('resultado', 'N/A')),
                    str(p.get('goles_local', 'N/A')),
                    str(p.get('goles_rival', 'N/A')),
                    str(p.get('ubicacion', 'N/A'))
                ])

        # Hoja 3: Asistencias
        ws_asistencias = wb.create_sheet("Asistencias")
        if not data["asistencias"]:
            ws_asistencias.append(["Sin datos registrados."])
        else:
            style_header(ws_asistencias, ["Fecha", "Jugador", "Evento", "Presente", "Justificado", "Motivo"])
            for a in data["asistencias"]:
                jug = a.get('jugador_nombre_completo', a.get('jugador', 'N/A'))
                evento_info = a.get('evento', 'N/A')
                ws_asistencias.append([
                    str(a.get('fecha', 'N/A')), 
                    str(jug), 
                    str(evento_info),
                    str(a.get('presente', a.get('estado', 'N/A'))), 
                    str(a.get('justificado', 'N/A')),
                    str(a.get('motivo_ausencia', a.get('motivo', 'N/A')))
                ])
                
        # Hoja 4: Estadísticas
        ws_estadisticas = wb.create_sheet("Estadísticas")
        if not data["estadisticas"]:
            ws_estadisticas.append(["Sin datos registrados."])
        else:
            style_header(ws_estadisticas, ["Jugador", "Partido", "Goles", "Asistencias", "Tarjetas", "Valoración", "Minutos"])
            for e in data["estadisticas"]:
                jug = e.get('jugador_nombre_completo', e.get('jugador', 'N/A'))
                tarjetas = f"A:{e.get('tarjetas_amarillas', 0)} R:{e.get('tarjetas_rojas', 0)}"
                ws_estadisticas.append([
                    str(jug), 
                    str(e.get('partido', 'N/A')),
                    str(e.get('goles', '0')), 
                    str(e.get('asistencias', '0')), 
                    tarjetas,
                    str(e.get('valoracion', 'N/A')),
                    str(e.get('minutos_jugados', 'N/A'))
                ])

        # Hoja 5: Evolución física
        ws_evolucion = wb.create_sheet("Evolución física")
        if not data["evolucion_fisica"]:
            ws_evolucion.append(["Sin datos registrados."])
        else:
            style_header(ws_evolucion, ["Fecha", "Jugador", "Peso kg", "Altura cm", "IMC", "Velocidad 40m"])
            for e in data["evolucion_fisica"]:
                jug = e.get('jugador_nombre_completo', e.get('jugador', 'N/A'))
                peso = e.get('peso_kg', e.get('peso', 0))
                altura = e.get('altura_cm', e.get('altura', 0))
                imc = 'N/A'
                try:
                    if float(altura) > 0 and float(peso) > 0:
                        imc = round(float(peso) / ((float(altura)/100)**2), 2)
                except (ValueError, TypeError):
                    pass

                ws_evolucion.append([
                    str(e.get('fecha', e.get('fecha_medicion', 'N/A'))), 
                    str(jug), 
                    str(peso), 
                    str(altura), 
                    str(imc), 
                    str(e.get('velocidad_40m', 'N/A'))
                ])

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=reporte-estadisticas-fecha.xlsx'
        wb.save(response)
        return response

class JugadoresPorEquipoView(APIView):
    def get(self, request):
        equipo_id = request.query_params.get('equipo')
        if not equipo_id:
            return Response({"error": "Debes especificar un equipo."}, status=status.HTTP_400_BAD_REQUEST)
        
        from .models import JugadorEquipo, Jugador
        from .serializers import JugadorSerializer
        
        jugador_ids = JugadorEquipo.objects.filter(equipo_id=equipo_id, activo=True).values_list('jugador_id', flat=True)
        jugadores = Jugador.objects.filter(id__in=jugador_ids)
        return Response(JugadorSerializer(jugadores, many=True).data)

